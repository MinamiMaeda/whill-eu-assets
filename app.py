"""
WHILL EU Asset Management — Cloud Version
Runs on Render.com | Database: Supabase PostgreSQL | Files: Supabase Storage
All config comes from environment variables (set on Render dashboard).
"""
from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
import os, datetime, csv, io, decimal, smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# ── Config: environment variables on Render, fallback to config.py locally ──
def _env(key, default=None):
    val = os.environ.get(key)
    if val is not None:
        return val
    try:
        import config
        return getattr(config, key, default)
    except ImportError:
        return default

DATABASE_URL       = _env("DATABASE_URL")
SUPABASE_URL       = _env("SUPABASE_URL")       # e.g. https://xxxx.supabase.co
SUPABASE_KEY       = _env("SUPABASE_SERVICE_KEY") # service_role key from Supabase
app.secret_key     = _env("SECRET_KEY", "whill-asset-mgmt-2025")
PORT               = int(_env("PORT", 5000))

EMAIL_ENABLED      = _env("EMAIL_ENABLED", "false").lower() == "true"
EMAIL_TO           = [e.strip() for e in _env("EMAIL_TO", "").split(",") if e.strip()]
EMAIL_FROM         = _env("EMAIL_FROM", "")
EMAIL_USERNAME     = _env("EMAIL_USERNAME", "")
EMAIL_PASSWORD     = _env("EMAIL_PASSWORD", "")
EMAIL_SMTP_HOST    = _env("EMAIL_SMTP_HOST", "smtp.gmail.com")
EMAIL_SMTP_PORT    = int(_env("EMAIL_SMTP_PORT", 587))

APP_URL            = _env("APP_URL", f"http://localhost:{PORT}")  # public URL for email links

STORAGE_BUCKET     = "asset-documents"
ALLOWED_EXT        = {"pdf", "png", "jpg", "jpeg", "gif", "docx", "xlsx", "msg", "zip"}

USERS     = ["Minami", "Sachiko", "Yuki", "Lo", "Other"]
APPROVERS = ["Yuki", "Lo"]

ASSET_TYPES = ["Demo Unit", "AD Device", "Laptop", "Display / Exhibit", "Other"]
DEP_METHODS = ["Declining Balance", "Straight-Line", "None"]
STATUSES    = ["Active", "With Customer", "In Storage", "In Transit", "Under Repair", "Sold", "Disposed"]
LOCATIONS   = ["NEN", "Specificlog", "France", "Germany", "Netherlands",
               "Italy", "UK", "Spain", "Belgium", "Poland", "Other"]
CURRENCIES  = ["EUR", "GBP", "USD", "JPY"]
TX_TYPES    = ["Sale", "Disposal", "Write-Off"]
DOC_TYPES   = ["Demo Contract", "Purchase Invoice", "Photo", "Maintenance Record", "Warranty", "Other"]

# ── Supabase Storage client ───────────────────────────────────
_storage = None
def get_storage():
    global _storage
    if _storage is None and SUPABASE_URL and SUPABASE_KEY:
        try:
            from supabase import create_client
            client = create_client(SUPABASE_URL, SUPABASE_KEY)
            _storage = client.storage
            # Ensure bucket exists
            try:
                _storage.create_bucket(STORAGE_BUCKET, options={"public": False})
            except Exception:
                pass  # Bucket already exists
        except Exception as e:
            print(f"⚠️  Supabase Storage init failed: {e}")
    return _storage

def upload_to_storage(file_bytes, storage_path, content_type="application/octet-stream"):
    """Upload file bytes to Supabase Storage. Returns public URL or None."""
    storage = get_storage()
    if not storage:
        return None
    try:
        storage.from_(STORAGE_BUCKET).upload(
            path=storage_path,
            file=file_bytes,
            file_options={"content-type": content_type, "upsert": "true"}
        )
        # Get signed URL valid for 10 years (effectively permanent for our use)
        res = storage.from_(STORAGE_BUCKET).create_signed_url(storage_path, 315360000)
        return res.get("signedURL") or res.get("signed_url")
    except Exception as e:
        print(f"⚠️  Storage upload failed: {e}")
        return None

def get_file_url(storage_path):
    """Get a fresh signed URL for a stored file."""
    if not storage_path:
        return None
    storage = get_storage()
    if not storage:
        return None
    try:
        res = storage.from_(STORAGE_BUCKET).create_signed_url(storage_path, 3600)
        return res.get("signedURL") or res.get("signed_url")
    except Exception as e:
        print(f"⚠️  Could not get file URL: {e}")
        return None

# ── DB helpers ────────────────────────────────────────────────
def get_db():
    return psycopg2.connect(DATABASE_URL)

def normalize(row):
    """Convert Decimal→float and datetime→ISO string so templates never crash."""
    if not row:
        return row
    out = {}
    for k, v in row.items():
        if isinstance(v, decimal.Decimal):
            out[k] = float(v)
        elif isinstance(v, (datetime.datetime, datetime.date)):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out

def fetchall(sql, params=None):
    conn = get_db()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or ())
        return [normalize(dict(r)) for r in cur.fetchall()]
    finally:
        conn.close()

def fetchone(sql, params=None):
    conn = get_db()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or ())
        row = cur.fetchone()
        return normalize(dict(row)) if row else None
    finally:
        conn.close()

def execute(sql, params=None):
    conn = get_db()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or ())
        conn.commit()
        try:    return normalize(dict(cur.fetchone())) if cur.description else None
        except: return None
    finally:
        conn.close()

def init_db():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS assets (
        id                   SERIAL PRIMARY KEY,
        asset_id             TEXT UNIQUE NOT NULL,
        serial_number        TEXT NOT NULL,
        asset_name           TEXT NOT NULL,
        asset_type           TEXT,
        model                TEXT,
        purchase_date        TEXT,
        purchase_value       NUMERIC,
        currency             TEXT DEFAULT 'EUR',
        dep_method           TEXT DEFAULT 'Declining Balance',
        useful_life_months   INTEGER DEFAULT 60,
        current_location     TEXT,
        status               TEXT DEFAULT 'Active',
        responsible          TEXT,
        notes                TEXT,
        approval_status      TEXT DEFAULT 'approved',
        created_at           TIMESTAMP DEFAULT NOW(),
        updated_at           TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS location_history (
        id          SERIAL PRIMARY KEY,
        asset_id    TEXT NOT NULL REFERENCES assets(asset_id),
        date_from   TEXT,
        date_to     TEXT,
        location    TEXT,
        country     TEXT,
        customer    TEXT,
        purpose     TEXT,
        shipped_by  TEXT,
        notes       TEXT,
        created_at  TIMESTAMP DEFAULT NOW(),
        created_by  TEXT
    );
    CREATE TABLE IF NOT EXISTS sales_disposal (
        id               SERIAL PRIMARY KEY,
        asset_id         TEXT NOT NULL REFERENCES assets(asset_id),
        tx_type          TEXT,
        tx_date          TEXT,
        book_value_at_tx NUMERIC,
        sale_price       NUMERIC,
        buyer            TEXT,
        buyer_contact    TEXT,
        invoice_ref      TEXT,
        notes            TEXT,
        approval_status  TEXT DEFAULT 'draft',
        approved_by      TEXT,
        approved_at      TIMESTAMP,
        reject_reason    TEXT,
        created_at       TIMESTAMP DEFAULT NOW(),
        created_by       TEXT
    );
    CREATE TABLE IF NOT EXISTS documents (
        id            SERIAL PRIMARY KEY,
        asset_id      TEXT NOT NULL REFERENCES assets(asset_id),
        doc_type      TEXT,
        doc_title     TEXT,
        doc_date      TEXT,
        file_path     TEXT,
        storage_path  TEXT,
        description   TEXT,
        uploaded_by   TEXT,
        created_at    TIMESTAMP DEFAULT NOW()
    );
    """)
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM assets")
    if cur.fetchone()[0] == 0:
        _seed(conn, cur)
    conn.close()

def _seed(conn, cur):
    assets = [
        ("DU-001","SN-WHILL-0001","Model C Demo Unit 1","Demo Unit","WHILL Model C","2022-04-01",8500,"EUR","Declining Balance",60,"Germany","With Customer","Minami","Deployed to Munich customer","approved"),
        ("DU-002","SN-WHILL-0002","Model C Demo Unit 2","Demo Unit","WHILL Model C","2022-04-01",8500,"EUR","Declining Balance",60,"France","With Customer","Minami","","approved"),
        ("DU-003","SN-WHILL-0003","Model C Demo Unit 3","Demo Unit","WHILL Model C","2022-07-15",8500,"EUR","Declining Balance",60,"NEN","In Storage","Minami","","approved"),
        ("DU-004","SN-WHILL-0004","Model C Lite Demo 1","Demo Unit","WHILL Model C Lite","2023-01-10",7200,"EUR","Declining Balance",60,"Belgium","With Customer","Minami","","approved"),
        ("DU-005","SN-WHILL-0005","Model C Lite Demo 2","Demo Unit","WHILL Model C Lite","2023-03-01",7200,"EUR","Declining Balance",60,"Netherlands","Active","Minami","","approved"),
        ("AD-001","SN-AD-0001","AD Device 1","AD Device","AD Model A","2021-06-01",3200,"EUR","Straight-Line",48,"NEN","Active","Sachiko","","approved"),
        ("AD-002","SN-AD-0002","AD Device 2","AD Device","AD Model A","2021-06-01",3200,"EUR","Straight-Line",48,"Germany","With Customer","Sachiko","","approved"),
        ("LT-001","SN-LT-0001","Laptop - Dell XPS","Laptop","Dell XPS 15","2022-09-01",1800,"EUR","Straight-Line",48,"NEN","Active","Minami","","approved"),
        ("EX-001","SN-EX-0001","Exhibition Display Stand","Display / Exhibit","Custom Stand","2022-01-15",2400,"EUR","Straight-Line",60,"NEN","In Storage","Minami","Exhibition unit","approved"),
    ]
    cur.executemany("""INSERT INTO assets (asset_id,serial_number,asset_name,asset_type,model,
        purchase_date,purchase_value,currency,dep_method,useful_life_months,
        current_location,status,responsible,notes,approval_status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (asset_id) DO NOTHING""", assets)
    conn.commit()

# ── Depreciation ──────────────────────────────────────────────
def calc_depreciation(asset):
    pv   = float(asset.get("purchase_value") or 0)
    pd   = str(asset.get("purchase_date") or "")[:10]
    meth = asset.get("dep_method") or "None"
    life = int(asset.get("useful_life_months") or 60)
    if not pd or pv == 0 or meth == "None":
        return {"monthly":0.0,"accumulated":0.0,"book_value":pv,"months_elapsed":0,"fully_dep":False}
    try:
        start = datetime.date.fromisoformat(pd)
    except:
        return {"monthly":0.0,"accumulated":0.0,"book_value":pv,"months_elapsed":0,"fully_dep":False}
    today  = datetime.date.today()
    months = max(0,(today.year-start.year)*12+(today.month-start.month))
    if meth == "Straight-Line":
        monthly=pv/life; accumulated=min(pv,monthly*months); book_value=max(0.0,pv-accumulated)
    elif meth == "Declining Balance":
        mr=(2/life)/12; book_value=pv*((1-mr)**months); accumulated=pv-book_value; monthly=book_value*mr
    else:
        monthly=accumulated=0.0; book_value=pv
    return {"monthly":round(float(monthly),2),"accumulated":round(float(accumulated),2),
            "book_value":round(float(book_value),2),"months_elapsed":months,"fully_dep":months>=life}

def allowed_file(f): return "." in f and f.rsplit(".",1)[1].lower() in ALLOWED_EXT
def require_login():
    if "user" not in session: return redirect(url_for("login"))
def is_approver(): return session.get("user") in APPROVERS

# ── Email ─────────────────────────────────────────────────────
def send_approval_email(subject, body):
    if not EMAIL_ENABLED:
        return
    try:
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = ", ".join(EMAIL_TO)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))
        with smtplib.SMTP(EMAIL_SMTP_HOST, EMAIL_SMTP_PORT) as s:
            s.starttls()
            s.login(EMAIL_USERNAME, EMAIL_PASSWORD)
            s.send_message(msg)
    except Exception as e:
        print(f"⚠️  Email failed: {e}")

# ══════════════════════════════════════════════════
# AUTH
# ══════════════════════════════════════════════════
@app.route("/", methods=["GET","POST"])
def login():
    if request.method == "POST":
        user = request.form.get("user","").strip()
        if user:
            session["user"] = user
            return redirect(url_for("dashboard"))
        flash("Please select your name.", "error")
    return render_template("login.html", users=USERS)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

# ══════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════
@app.route("/dashboard")
def dashboard():
    r = require_login()
    if r: return r
    all_assets     = fetchall("SELECT * FROM assets")
    pending_assets = fetchall("SELECT * FROM assets WHERE approval_status='pending' ORDER BY created_at DESC")
    pending_txns   = fetchall("""
        SELECT sd.*, a.asset_name, a.serial_number, a.asset_type, a.purchase_value
        FROM sales_disposal sd JOIN assets a ON sd.asset_id=a.asset_id
        WHERE sd.approval_status='pending' ORDER BY sd.created_at DESC""")
    approved = [a for a in all_assets if a["approval_status"]=="approved"]
    for a in approved:
        a.update(calc_depreciation(a))
    for t in pending_txns:
        ax = next((x for x in all_assets if x["asset_id"]==t["asset_id"]), None)
        if ax: t["book_value_now"] = calc_depreciation(ax)["book_value"]
    total_purchase = sum(float(a["purchase_value"] or 0) for a in approved)
    total_book     = sum(float(a["book_value"]) for a in approved)
    total_monthly  = sum(float(a["monthly"]) for a in approved if a["status"] not in ("Sold","Disposed"))
    by_type={}; by_status={}; by_loc={}
    for a in approved:
        by_type[a["asset_type"]]      = by_type.get(a["asset_type"],0)+1
        by_status[a["status"]]        = by_status.get(a["status"],0)+1
        by_loc[a["current_location"]] = by_loc.get(a["current_location"],0)+1
    return render_template("dashboard.html",
        assets=approved, total_purchase=total_purchase, total_book=total_book,
        total_monthly=total_monthly, by_type=by_type, by_status=by_status, by_loc=by_loc,
        pending_assets=pending_assets, pending_txns=pending_txns,
        is_approver=is_approver(), today=datetime.date.today().strftime("%d %B %Y"))

# ── Approval actions ──────────────────────────────
@app.route("/approve/asset/<int:aid>", methods=["POST"])
def approve_asset(aid):
    r = require_login()
    if r: return r
    if not is_approver(): flash("Only Yuki or Lo can approve.", "error"); return redirect(url_for("dashboard"))
    action = request.form.get("action"); reason = request.form.get("reason","")
    asset  = fetchone("SELECT * FROM assets WHERE id=%s", (aid,))
    if action == "approve":
        execute("UPDATE assets SET approval_status='approved',status='Active' WHERE id=%s", (aid,))
        execute("""INSERT INTO location_history (asset_id,date_from,location,country,purpose,created_by)
            VALUES (%s,%s,%s,'','Initial receipt',%s)""",
            (asset["asset_id"],asset.get("purchase_date",""),asset.get("current_location",""),session["user"]))
        flash(f"✅ Asset {asset['asset_id']} approved.", "success")
    else:
        execute("UPDATE assets SET approval_status='rejected',notes=%s WHERE id=%s",
            (f"[REJECTED: {reason}] "+(asset.get("notes") or ""), aid))
        flash(f"❌ Asset {asset['asset_id']} rejected.", "error")
    return redirect(url_for("dashboard"))

@app.route("/approve/transaction/<int:tid>", methods=["POST"])
def approve_transaction(tid):
    r = require_login()
    if r: return r
    if not is_approver(): flash("Only Yuki or Lo can approve.", "error"); return redirect(url_for("dashboard"))
    action = request.form.get("action"); reason = request.form.get("reason","")
    tx = fetchone("SELECT * FROM sales_disposal WHERE id=%s", (tid,))
    if action == "approve":
        execute("UPDATE sales_disposal SET approval_status='approved',approved_by=%s,approved_at=NOW() WHERE id=%s",
            (session["user"], tid))
        execute("UPDATE assets SET status=%s,updated_at=NOW() WHERE asset_id=%s",
            ("Sold" if tx["tx_type"]=="Sale" else "Disposed", tx["asset_id"]))
        flash(f"✅ {tx['tx_type']} for {tx['asset_id']} approved.", "success")
    else:
        execute("UPDATE sales_disposal SET approval_status='rejected',reject_reason=%s,approved_by=%s,approved_at=NOW() WHERE id=%s",
            (reason, session["user"], tid))
        flash(f"❌ {tx['tx_type']} for {tx['asset_id']} rejected.", "error")
    return redirect(url_for("dashboard"))

# ══════════════════════════════════════════════════
# ASSETS
# ══════════════════════════════════════════════════
@app.route("/assets")
def assets():
    r = require_login()
    if r: return r
    q=request.args.get("q","").strip(); f_type=request.args.get("type","")
    f_status=request.args.get("status",""); f_loc=request.args.get("location","")
    sql="SELECT * FROM assets WHERE approval_status NOT IN ('pending','rejected')"; params=[]
    if q:
        sql+=" AND (asset_id ILIKE %s OR serial_number ILIKE %s OR asset_name ILIKE %s OR model ILIKE %s)"
        params+=[f"%{q}%"]*4
    if f_type:   sql+=" AND asset_type=%s";       params.append(f_type)
    if f_status: sql+=" AND status=%s";            params.append(f_status)
    if f_loc:    sql+=" AND current_location=%s";  params.append(f_loc)
    sql+=" ORDER BY asset_id"
    rows = fetchall(sql, params)
    for a in rows: a.update(calc_depreciation(a))
    return render_template("assets.html", assets=rows, q=q, f_type=f_type,
        f_status=f_status, f_loc=f_loc, asset_types=ASSET_TYPES, statuses=STATUSES, locations=LOCATIONS)

@app.route("/assets/new", methods=["GET","POST"])
def asset_new():
    r = require_login()
    if r: return r
    if request.method == "POST":
        f = request.form
        send_approval = f.get("send_approval") == "yes"
        approval    = "pending" if send_approval else "approved"
        init_status = "Pending Approval" if send_approval else f.get("status","Active")
        execute("""INSERT INTO assets (asset_id,serial_number,asset_name,asset_type,model,
            purchase_date,purchase_value,currency,dep_method,useful_life_months,
            current_location,status,responsible,notes,approval_status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
            f["asset_id"],f["serial_number"],f["asset_name"],f["asset_type"],
            f.get("model",""),f.get("purchase_date",""),
            float(f["purchase_value"] or 0),f.get("currency","EUR"),
            f.get("dep_method","Declining Balance"),
            int(f.get("useful_life_months",60) or 60),
            f.get("current_location",""),init_status,
            f.get("responsible",session["user"]),f.get("notes",""),approval))
        if approval == "approved":
            execute("""INSERT INTO location_history (asset_id,date_from,location,country,purpose,created_by)
                VALUES (%s,%s,%s,'','Initial receipt',%s)""",
                (f["asset_id"],f.get("purchase_date",""),f.get("current_location",""),session["user"]))
            flash(f"✅ Asset {f['asset_id']} added.", "success")
            return redirect(url_for("asset_detail", asset_id=f["asset_id"]))
        else:
            send_approval_email(
                subject=f"[WHILL Assets] Approval Request — New Asset {f['asset_id']}",
                body=f"""<h3>New Asset Approval Request</h3>
                <p><b>Submitted by:</b> {session['user']}</p>
                <p><b>Asset ID:</b> {f['asset_id']} | <b>Serial:</b> {f['serial_number']}</p>
                <p><b>Name:</b> {f['asset_name']} | <b>Type:</b> {f['asset_type']}</p>
                <p><b>Value:</b> {f.get('currency','EUR')} {f.get('purchase_value','')}</p>
                <p><b>Location:</b> {f.get('current_location','')}</p>
                <p><a href="{APP_URL}/dashboard">→ Review on Dashboard</a></p>"""
            )
            flash(f"⏳ Asset {f['asset_id']} submitted for approval.", "success")
            return redirect(url_for("assets"))
    count = fetchone("SELECT COUNT(*) as c FROM assets")["c"]
    return render_template("asset_form.html", asset=None, mode="new",
        asset_types=ASSET_TYPES, dep_methods=DEP_METHODS, statuses=STATUSES,
        locations=LOCATIONS, currencies=CURRENCIES, users=USERS,
        next_id=f"DU-{int(count)+1:03d}", is_approver=is_approver())

@app.route("/assets/<asset_id>")
def asset_detail(asset_id):
    r = require_login()
    if r: return r
    asset = fetchone("SELECT * FROM assets WHERE asset_id=%s", (asset_id,))
    if not asset: flash("Asset not found.", "error"); return redirect(url_for("assets"))
    asset.update(calc_depreciation(asset))
    lh   = fetchall("SELECT * FROM location_history WHERE asset_id=%s ORDER BY date_from DESC", (asset_id,))
    docs = fetchall("SELECT * FROM documents WHERE asset_id=%s ORDER BY created_at DESC", (asset_id,))
    txns = fetchall("SELECT * FROM sales_disposal WHERE asset_id=%s ORDER BY tx_date DESC", (asset_id,))
    # Attach fresh signed URLs to each document
    for doc in docs:
        if doc.get("storage_path"):
            doc["download_url"] = get_file_url(doc["storage_path"])
        else:
            doc["download_url"] = None
    return render_template("asset_detail.html", asset=asset, lh=lh, docs=docs, txns=txns,
        locations=LOCATIONS, doc_types=DOC_TYPES, tx_types=TX_TYPES,
        users=USERS, is_approver=is_approver())

@app.route("/assets/<asset_id>/edit", methods=["GET","POST"])
def asset_edit(asset_id):
    r = require_login()
    if r: return r
    asset = fetchone("SELECT * FROM assets WHERE asset_id=%s", (asset_id,))
    if not asset: return redirect(url_for("assets"))
    if request.method == "POST":
        f=request.form
        execute("""UPDATE assets SET serial_number=%s,asset_name=%s,asset_type=%s,model=%s,
            purchase_date=%s,purchase_value=%s,currency=%s,dep_method=%s,useful_life_months=%s,
            current_location=%s,status=%s,responsible=%s,notes=%s,updated_at=NOW()
            WHERE asset_id=%s""", (
            f["serial_number"],f["asset_name"],f["asset_type"],f.get("model",""),
            f.get("purchase_date",""),float(f["purchase_value"] or 0),
            f.get("currency","EUR"),f.get("dep_method","Declining Balance"),
            int(f.get("useful_life_months",60) or 60),
            f.get("current_location",""),f.get("status","Active"),
            f.get("responsible",""),f.get("notes",""),asset_id))
        flash("Asset updated.", "success")
        return redirect(url_for("asset_detail", asset_id=asset_id))
    return render_template("asset_form.html", asset=asset, mode="edit",
        asset_types=ASSET_TYPES, dep_methods=DEP_METHODS, statuses=STATUSES,
        locations=LOCATIONS, currencies=CURRENCIES, users=USERS,
        next_id=None, is_approver=is_approver())

@app.route("/assets/<asset_id>/update-life", methods=["POST"])
def update_life(asset_id):
    r = require_login()
    if r: return r
    months = max(1, int(request.form.get("useful_life_months",60) or 60))
    execute("UPDATE assets SET useful_life_months=%s,updated_at=NOW() WHERE asset_id=%s", (months,asset_id))
    flash(f"Depreciation period updated to {months} months.", "success")
    return redirect(request.referrer or url_for("depreciation"))

@app.route("/assets/<asset_id>/location/add", methods=["POST"])
def location_add(asset_id):
    r = require_login()
    if r: return r
    f=request.form
    execute("UPDATE location_history SET date_to=%s WHERE asset_id=%s AND (date_to IS NULL OR date_to='')",
        (f.get("date_from",""), asset_id))
    execute("""INSERT INTO location_history
        (asset_id,date_from,date_to,location,country,customer,purpose,shipped_by,notes,created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
        asset_id,f.get("date_from",""),f.get("date_to",""),f.get("location",""),
        f.get("country",""),f.get("customer",""),f.get("purpose",""),
        f.get("shipped_by",""),f.get("notes",""),session["user"]))
    execute("UPDATE assets SET current_location=%s,status=%s,updated_at=NOW() WHERE asset_id=%s",
        (f.get("location",""),f.get("status","With Customer"),asset_id))
    flash("Location updated.", "success")
    return redirect(url_for("asset_detail", asset_id=asset_id))

# ── Sales & Disposal ──────────────────────────────
@app.route("/assets/<asset_id>/transaction/add", methods=["POST"])
def transaction_add(asset_id):
    r = require_login()
    if r: return r
    f=request.form
    send_approval = f.get("send_approval") == "yes"
    asset = fetchone("SELECT * FROM assets WHERE asset_id=%s", (asset_id,))
    dep   = calc_depreciation(asset)
    approval_status = "pending" if send_approval else "draft"
    execute("""INSERT INTO sales_disposal
        (asset_id,tx_type,tx_date,book_value_at_tx,sale_price,buyer,buyer_contact,
         invoice_ref,notes,approval_status,created_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
        asset_id,f.get("tx_type","Sale"),f.get("tx_date",""),
        dep["book_value"],float(f.get("sale_price",0) or 0),
        f.get("buyer",""),f.get("buyer_contact",""),
        f.get("invoice_ref",""),f.get("notes",""),approval_status,session["user"]))
    if send_approval:
        send_approval_email(
            subject=f"[WHILL Assets] Approval Request — {f.get('tx_type')} for {asset_id}",
            body=f"""<h3>{f.get('tx_type')} Approval Request</h3>
            <p><b>Submitted by:</b> {session['user']}</p>
            <p><b>Asset:</b> {asset_id} — {asset.get('asset_name','')}</p>
            <p><b>Book Value:</b> €{dep['book_value']:,.2f} | <b>Sale Price:</b> €{float(f.get('sale_price',0) or 0):,.2f}</p>
            <p><a href="{APP_URL}/dashboard">→ Review on Dashboard</a></p>"""
        )
        flash(f"⏳ {f.get('tx_type')} submitted for approval.", "success")
    else:
        flash(f"💾 {f.get('tx_type')} saved as draft.", "success")
    return redirect(url_for("asset_detail", asset_id=asset_id))

@app.route("/transaction/<int:tid>/submit", methods=["POST"])
def transaction_submit(tid):
    r = require_login()
    if r: return r
    tx = fetchone("SELECT * FROM sales_disposal WHERE id=%s", (tid,))
    if not tx: flash("Transaction not found.", "error"); return redirect(url_for("transactions"))
    execute("UPDATE sales_disposal SET approval_status='pending' WHERE id=%s", (tid,))
    asset = fetchone("SELECT * FROM assets WHERE asset_id=%s", (tx["asset_id"],))
    send_approval_email(
        subject=f"[WHILL Assets] Approval Request — {tx.get('tx_type')} for {tx['asset_id']}",
        body=f"""<h3>{tx.get('tx_type')} Approval Request</h3>
        <p><b>Asset:</b> {tx['asset_id']} — {asset.get('asset_name','') if asset else ''}</p>
        <p><b>Book Value:</b> €{float(tx.get('book_value_at_tx') or 0):,.2f}</p>
        <p><a href="{APP_URL}/dashboard">→ Review on Dashboard</a></p>"""
    )
    flash("⏳ Submitted for approval.", "success")
    return redirect(url_for("asset_detail", asset_id=tx["asset_id"]))

# ── Document upload → Supabase Storage ───────────
@app.route("/assets/<asset_id>/document/add", methods=["POST"])
def document_add(asset_id):
    r = require_login()
    if r: return r
    f=request.form; file=request.files.get("file"); storage_path=""
    if file and file.filename and allowed_file(file.filename):
        asset  = fetchone("SELECT serial_number FROM assets WHERE asset_id=%s", (asset_id,))
        serial = secure_filename(asset["serial_number"] if asset else asset_id)
        ts     = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        fn     = secure_filename(file.filename)
        storage_path = f"{serial}/{ts}_{fn}"
        content_type = file.content_type or "application/octet-stream"
        file_bytes = file.read()
        result = upload_to_storage(file_bytes, storage_path, content_type)
        if not result:
            flash("⚠️  File upload failed — document record saved without attachment.", "error")
            storage_path = ""
    execute("""INSERT INTO documents (asset_id,doc_type,doc_title,doc_date,storage_path,description,uploaded_by)
        VALUES (%s,%s,%s,%s,%s,%s,%s)""", (
        asset_id,f.get("doc_type","Other"),f.get("doc_title",""),
        f.get("doc_date",""),storage_path,f.get("description",""),session["user"]))
    if storage_path:
        flash("✅ Document uploaded.", "success")
    return redirect(url_for("asset_detail", asset_id=asset_id))

@app.route("/documents/<int:doc_id>/download")
def download_document(doc_id):
    r = require_login()
    if r: return r
    doc = fetchone("SELECT * FROM documents WHERE id=%s", (doc_id,))
    if not doc or not doc.get("storage_path"):
        flash("Document not found.", "error")
        return redirect(url_for("assets"))
    url = get_file_url(doc["storage_path"])
    if url:
        return redirect(url)
    flash("Could not retrieve file.", "error")
    return redirect(url_for("assets"))

# ══════════════════════════════════════════════════
# DEPRECIATION
# ══════════════════════════════════════════════════
@app.route("/depreciation")
def depreciation():
    r = require_login()
    if r: return r
    f_type=request.args.get("type","")
    sql="SELECT * FROM assets WHERE status NOT IN ('Sold','Disposed') AND approval_status='approved'"
    params=[]
    if f_type: sql+=" AND asset_type=%s"; params.append(f_type)
    sql+=" ORDER BY asset_type, asset_id"
    rows = fetchall(sql, params)
    for a in rows: a.update(calc_depreciation(a))
    return render_template("depreciation.html", assets=rows,
        total_monthly=sum(float(a["monthly"]) for a in rows),
        total_book=sum(float(a["book_value"]) for a in rows),
        total_purchase=sum(float(a["purchase_value"] or 0) for a in rows),
        f_type=f_type, asset_types=ASSET_TYPES)

# ══════════════════════════════════════════════════
# TRANSACTIONS
# ══════════════════════════════════════════════════
@app.route("/transactions")
def transactions():
    r = require_login()
    if r: return r
    rows = fetchall("""SELECT sd.*, a.asset_name, a.serial_number, a.asset_type, a.purchase_value
        FROM sales_disposal sd JOIN assets a ON sd.asset_id=a.asset_id
        ORDER BY sd.tx_date DESC NULLS LAST""")
    for row in rows:
        sp=float(row["sale_price"] or 0); bv=float(row["book_value_at_tx"] or 0)
        row["gain_loss"]=round(sp-bv,2) if row["tx_type"]=="Sale" else None
    return render_template("transactions.html", transactions=rows, is_approver=is_approver())

# ══════════════════════════════════════════════════
# LOCATION HISTORY
# ══════════════════════════════════════════════════
@app.route("/location-history")
def location_history():
    r = require_login()
    if r: return r
    q=request.args.get("q","").strip()
    if q:
        rows=fetchall("""SELECT lh.*,a.asset_name,a.asset_type FROM location_history lh
            JOIN assets a ON lh.asset_id=a.asset_id
            WHERE lh.asset_id ILIKE %s OR lh.location ILIKE %s OR lh.customer ILIKE %s
            ORDER BY lh.date_from DESC""",[f"%{q}%"]*3)
    else:
        rows=fetchall("""SELECT lh.*,a.asset_name,a.asset_type FROM location_history lh
            JOIN assets a ON lh.asset_id=a.asset_id ORDER BY lh.date_from DESC""")
    return render_template("location_history.html", history=rows, q=q)

# ══════════════════════════════════════════════════
# EXPORTS
# ══════════════════════════════════════════════════
@app.route("/export/assets/csv")
def export_assets_csv():
    r = require_login()
    if r: return r
    rows=fetchall("SELECT * FROM assets WHERE approval_status='approved' ORDER BY asset_id")
    for a in rows:
        d=calc_depreciation(a); a["monthly_dep"]=d["monthly"]; a["book_value"]=d["book_value"]; a["accumulated"]=d["accumulated"]
    output=io.StringIO()
    fields=["asset_id","serial_number","asset_name","asset_type","model","purchase_date",
            "purchase_value","currency","dep_method","useful_life_months","current_location",
            "status","responsible","monthly_dep","book_value","accumulated","notes"]
    w=csv.DictWriter(output,fieldnames=fields); w.writeheader()
    for a in rows: w.writerow({k:a.get(k,"") for k in fields})
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype="text/csv",
        as_attachment=True, download_name=f"WHILL_Assets_{datetime.date.today()}.csv")

@app.route("/export/assets/xlsx")
def export_assets_xlsx():
    r = require_login()
    if r: return r
    rows=fetchall("SELECT * FROM assets WHERE approval_status='approved' ORDER BY asset_id")
    for a in rows:
        d=calc_depreciation(a); a["monthly_dep"]=d["monthly"]; a["book_value"]=d["book_value"]; a["accumulated"]=d["accumulated"]
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Asset Register"
    hdrs=["Asset ID","Serial Number","Asset Name","Type","Model","Purchase Date",
          "Purchase Value","Currency","Dep. Method","Useful Life (mo)","Location",
          "Status","Responsible","Monthly Dep.","Book Value","Accumulated Dep.","Notes"]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
        c.fill=PatternFill("solid",fgColor="1F7BC1")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    keys=["asset_id","serial_number","asset_name","asset_type","model","purchase_date",
          "purchase_value","currency","dep_method","useful_life_months","current_location",
          "status","responsible","monthly_dep","book_value","accumulated","notes"]
    for ri,a in enumerate(rows,2):
        for col,k in enumerate(keys,1): ws.cell(row=ri,column=col,value=a.get(k,""))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,download_name=f"WHILL_Assets_{datetime.date.today()}.xlsx")

@app.route("/export/depreciation/xlsx")
def export_depreciation_xlsx():
    r = require_login()
    if r: return r
    rows=fetchall("SELECT * FROM assets WHERE status NOT IN ('Sold','Disposed') AND approval_status='approved' ORDER BY asset_type,asset_id")
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Depreciation"
    hdrs=["Asset ID","Serial Number","Asset Name","Type","Purchase Date","Purchase Value",
          "Method","Useful Life (mo)","Months Elapsed","Monthly Dep.","Accumulated Dep.","Book Value","Fully Dep."]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=1,column=col,value=h)
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
        c.fill=PatternFill("solid",fgColor="1F7BC1")
        c.alignment=Alignment(horizontal="center",vertical="center")
    for ri,a in enumerate(rows,2):
        d=calc_depreciation(a)
        for col,val in enumerate([a["asset_id"],a["serial_number"],a["asset_name"],a["asset_type"],
            a["purchase_date"],a["purchase_value"],a["dep_method"],a["useful_life_months"],
            d["months_elapsed"],d["monthly"],d["accumulated"],d["book_value"],
            "Yes" if d["fully_dep"] else "No"],1):
            ws.cell(row=ri,column=col,value=val)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,download_name=f"WHILL_Depreciation_{datetime.date.today()}.xlsx")

# ── Startup ───────────────────────────────────────
if __name__ == "__main__":
    print("\n" + "="*52)
    print("  WHILL Asset Management — Cloud (Render)")
    print("  Connecting to Supabase database...")
    try:
        init_db()
        print("  ✅ Database ready")
    except Exception as e:
        print(f"  ❌ Database failed: {e}"); exit(1)
    storage = get_storage()
    print(f"  {'✅ Supabase Storage ready' if storage else '⚠️  Storage not configured (uploads disabled)'}")
    print(f"  {'✅ Email ON → ' + str(EMAIL_TO) if EMAIL_ENABLED else 'ℹ️  Email OFF'}")
    print(f"  Open: http://localhost:{PORT}")
    print("="*52 + "\n")
    app.run(debug=False, host="0.0.0.0", port=PORT)
